import string
import time
from dataclasses import dataclass
from datetime import timedelta, datetime
from itertools import groupby

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

hora_inicio_jornada = timedelta(hours=10, minutes=0, seconds=0)


@dataclass
class DadosCliente:
    subtipo_prazo: string
    adverso: string
    processo_judicial: string
    data_hora_prazo: string
    tempo_execucao: int = 0
    hora_lancamento_timesheet: timedelta = timedelta(hours=0, minutes=0, seconds=0)


def determinar_tempo_prazo(subtipo_prazo) -> int:
    if subtipo_prazo == 'Regularizar Representação':
        return 7
    elif 'Contrarrazoes' in subtipo_prazo:
        return 20
    elif subtipo_prazo == 'Apresentar Documentos':
        return 6
    elif subtipo_prazo == 'Recurso Ordinário':
        return 60
    else:
        return 5


def fazerLogin():
    options = Options()
    options.add_experimental_option("detach", True)
    browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    browser.get("https://autuori.elawio.com.br/timesheet/CadastrarPeriodo")
    browser.maximize_window()

    email = browser.find_element(By.ID, "Email")

    email.send_keys("leticia.oliveira@autuori.com.br")

    senha = browser.find_element(By.ID, "Password")

    senha.send_keys("123*Autuori!")

    login = browser.find_element(By.XPATH, value="/html/body/div/div/form/table/tbody/tr[5]/td/button")

    login.click()

    return browser


def preencherTimeSheet(
        browser,
        _data_inicio,
        _data_fim,
        _hora_inicio,
        _hora_fim,
        _descricao_abertura,
        _descricao_conclusao,
        _atividade="Elaboração de prazo"

):
    time.sleep(15)

    timeSheetList = browser.find_element(By.XPATH, value="/html/body/div[6]/div[1]/div/ul/li[13]")
    timeSheetList.click()

    time.sleep(5)

    horas_por_periodo = browser.find_element(By.XPATH, value="/html/body/div[6]/div[1]/div/ul/li[13]/ul/li[5]")
    horas_por_periodo.click()

    time.sleep(5)

    atividade = browser.find_element(By.XPATH,
                                     "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[2]/div/div/button/div/div/div")
    atividade.click()

    time.sleep(2)

    elaboracao_prazo = browser.find_element(By.XPATH,
                                            "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[2]/div/div/div/div[2]/ul/li[58]")
    elaboracao_prazo.click()

    data_inicio = browser.find_element(By.XPATH,
                                       "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[3]/div[1]/div/input")
    data_inicio.send_keys(_data_inicio)

    data_fim = browser.find_element(By.XPATH,
                                    "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[4]/div[1]/div/input")
    data_fim.send_keys(_data_fim)

    hora_inicio = browser.find_element(By.XPATH,
                                       "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[3]/div[2]/div/input")
    hora_inicio.send_keys(_hora_inicio)

    hora_fim = browser.find_element(By.XPATH,
                                    "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[4]/div[2]/div/input")
    hora_fim.send_keys(_hora_fim)

    descricao_abertura = browser.find_element(By.XPATH,
                                              "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[5]/div/textarea")
    descricao_abertura.send_keys(_descricao_abertura)

    descricao_conclusao = browser.find_element(By.XPATH,
                                               "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[6]/div/textarea")
    descricao_conclusao.send_keys(_descricao_conclusao)

    cliente = browser.find_element(By.XPATH,
                                   "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/form/div[7]/div[1]/label[2]")
    cliente.click()

    time.sleep(2)

    nome_cliente = browser.find_element(By.XPATH,
                                        "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div[2]/form/div[1]/input")
    nome_cliente.send_keys("carrefour")

    pesquisar = browser.find_element(By.XPATH,
                                     "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div[2]/form/button")
    pesquisar.click()

    time.sleep(3)

    click_carrefour_client = browser.find_element(By.XPATH,
                                                  "/html/body/div[6]/div[2]/div/div[1]/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div[2]/form/div[2]/div/div/table/tbody/tr[1]/td[1]")
    click_carrefour_client.click()

    salvar = browser.find_element(By.XPATH,
                                  "/html/body/div[6]/div[2]/div/div[1]/div[6]/a[1]")
    salvar.click()

    time.sleep(5)

    browser.quit()


def formatar_horario(ultimo_horario):
    horario = str(ultimo_horario)[:5]
    if horario[2] == ":":
        return horario
    else:
        return f'0{horario[:4]}'


def cronograma_dia():
    workbook = openpyxl.load_workbook('compromissos.xlsx')
    sheet_compromissos = workbook['CompromissoViewModel']

    list_dados = []

    # cria objetos com infos dos clientes

    for linha in sheet_compromissos.iter_rows(min_row=2):
        tempo_execucao = determinar_tempo_prazo(linha[11].value)
        list_dados.append(
            DadosCliente(
                subtipo_prazo=linha[11].value,
                adverso=linha[12].value,
                processo_judicial=linha[14].value,
                data_hora_prazo=datetime.strptime(linha[5].value,
                                                  "%d/%m/%Y %H:%M:%S").strftime("%d/%m/%Y"),
                tempo_execucao=tempo_execucao
            )
        )

    # ordena e agrupa objetos com base no subtipo do prazo

    ordered_list_by_date = sorted(list_dados, key=lambda x: x.data_hora_prazo, reverse=False)

    grouped_by_date = [list(result) for key, result in groupby(
        ordered_list_by_date, key=lambda cliente: cliente.data_hora_prazo)]

    grouped_by_date_and_subtipo_prazo = {}

    for date_group in grouped_by_date:
        ordered_list_by_subtipo_prazo = sorted(date_group, key=lambda x: x.subtipo_prazo, reverse=False)
        grouped_by_subtipo_prazo = [list(result) for key, result in groupby(
            ordered_list_by_subtipo_prazo, key=lambda cliente: cliente.subtipo_prazo)]

        grouped_by_date_and_subtipo_prazo[date_group[0].data_hora_prazo] = grouped_by_subtipo_prazo

    # cria planilha com informacoes

    wb = openpyxl.Workbook()

    wb.create_sheet('dados')

    sheet = wb.active

    # cria cabecalho

    sheet.cell(row=1, column=1).value = 'Subtipo Prazo'
    sheet.cell(row=1, column=2).value = 'Adverso'
    sheet.cell(row=1, column=3).value = 'Processo Judicial'
    sheet.cell(row=1, column=4).value = 'Data Inicio'
    sheet.cell(row=1, column=5).value = 'Data Fim'
    sheet.cell(row=1, column=6).value = 'Data/Hora'

    # considera o inicio da jornada e soma os tempos conforme o tempo gasto por cada prazo e adiciona no workbook
    linha_atual = 2
    total_enviados = 0

    for date, list_prazos in grouped_by_date_and_subtipo_prazo.items():
        print(f'{date} -> {list_prazos}')
        ultimo_horario = hora_inicio_jornada
        prazo_atual_do_dia = 0
        for list_subtipo_prazo in list_prazos:
            for dado_cliente in list_subtipo_prazo:
                dado_cliente.data_inicio = ultimo_horario
                ultimo_horario = ultimo_horario + timedelta(hours=0, minutes=dado_cliente.tempo_execucao + 1, seconds=0)
                sheet.cell(row=linha_atual, column=1).value = dado_cliente.subtipo_prazo
                sheet.cell(row=linha_atual, column=2).value = dado_cliente.adverso
                sheet.cell(row=linha_atual, column=3).value = dado_cliente.processo_judicial
                sheet.cell(row=linha_atual, column=4).value = dado_cliente.data_inicio
                sheet.cell(row=linha_atual, column=5).value = ultimo_horario - timedelta(hours=0, minutes=1,
                                                                                         seconds=0)
                sheet.cell(row=linha_atual, column=6).value = dado_cliente.data_hora_prazo
                linha_atual = linha_atual + 1
                prazo_atual_do_dia = prazo_atual_do_dia + 1

                try:
                    browser = fazerLogin()
                    preencherTimeSheet(browser=browser,
                                       _data_inicio=dado_cliente.data_hora_prazo,
                                       _data_fim=dado_cliente.data_hora_prazo,
                                       _hora_inicio=formatar_horario(dado_cliente.data_inicio),
                                       _hora_fim=formatar_horario(ultimo_horario - timedelta(hours=0, minutes=1,
                                                                                             seconds=0)),
                                       _descricao_abertura=f'{dado_cliente.subtipo_prazo} {dado_cliente.adverso} {dado_cliente.processo_judicial}',
                                       _descricao_conclusao=f'{dado_cliente.subtipo_prazo} {dado_cliente.adverso} {dado_cliente.processo_judicial}'
                                       )
                except Exception as error:
                    print(f'Erro ao processar dados do cliente: {dado_cliente}')
                    continue

                total_enviados = total_enviados + 1
                print(
                    f'Quantidade de prazos enviados: {prazo_atual_do_dia} do dia {date}, total no arquivo: {len(list_dados)}, total enviados: {total_enviados}')
                time.sleep(5)

                # salva workbook

                wb.save('./dados_cliente.xlsx')


if __name__ == '__main__':
    cronograma_dia()
